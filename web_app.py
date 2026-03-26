from collections import OrderedDict
from io import BytesIO
from datetime import datetime, timezone
from pathlib import Path
from urllib.parse import quote
from uuid import uuid4
from zipfile import ZIP_DEFLATED, ZipFile

import pandas as pd
import uvicorn
from fastapi import FastAPI, HTTPException, Request
from fastapi.responses import JSONResponse, Response
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from openpyxl import Workbook
from openpyxl.styles import Font

from 共通スクリプト.Excel出力.excel_exporter import build_excel_bytes
from 共通スクリプト.analysis_service import (
    DEFAULT_ANALYSIS_KEYS,
    create_activity_case_drilldown,
    analyze_prepared_event_log,
    create_analysis_records,
    create_bottleneck_summary,
    create_case_trace_details,
    create_dashboard_summary,
    create_impact_summary,
    create_log_diagnostics,
    create_pattern_index_entries,
    create_rule_based_insights,
    create_root_cause_summary,
    create_transition_case_drilldown,
    filter_prepared_df,
    filter_prepared_df_by_pattern,
    get_filter_options,
    create_variant_flow_snapshot,
    create_variant_summary,
    create_pattern_flow_snapshot,
    create_pattern_bottleneck_details,
    get_available_analysis_definitions,
    load_prepared_event_log,
    merge_filter_params,
    normalize_filter_params,
    normalize_filter_column_settings,
)


BASE_DIR = Path(__file__).resolve().parent
SAMPLE_FILE = BASE_DIR / "sample_event_log.csv"
MAX_STORED_RUNS = 5
PREVIEW_ROW_COUNT = 10
PROCESS_FLOW_PATTERN_CAP = 300
MAX_PATTERN_FLOW_CACHE = 24
FILTER_PARAM_NAMES = (
    "date_from",
    "date_to",
    "filter_value_1",
    "filter_value_2",
    "filter_value_3",
    "activity_mode",
    "activity_values",
)
FILTER_COLUMN_NAMES = ("filter_column_1", "filter_column_2", "filter_column_3")
FILTER_LABEL_NAMES = ("filter_label_1", "filter_label_2", "filter_label_3")
REPORT_SHEET_NAMES = {
    "summary": "サマリー",
    "frequency": "頻度分析",
    "pattern": "処理順パターン分析",
    "variant": "Variant分析",
    "bottleneck": "ボトルネック分析",
    "impact": "改善インパクト分析",
    "drilldown": "ドリルダウン",
    "case_trace": "ケース追跡",
}
REPORT_HEADER_LABELS = {
    "run_id": "実行ID",
    "analysis_key": "分析種別",
    "analysis_name": "分析名",
    "source_file_name": "元ファイル名",
    "analysis_executed_at": "分析実行日時",
    "exported_at": "出力日時",
    "case_count": "対象ケース数",
    "event_count": "対象イベント数",
    "applied_filters": "適用フィルタ条件",
    "selected_variant": "選択中Variant",
    "selected_activity": "選択中Activity",
    "selected_transition": "選択中遷移",
    "selected_case_id": "選択中Case ID",
    "rank": "順位",
    "variant_id": "Variant ID",
    "count": "件数",
    "case_count": "対象ケース数",
    "ratio": "比率",
    "pattern": "パターン",
    "activity_count": "Activity数",
    "avg_case_duration": "平均所要時間",
    "avg_duration": "平均待ち時間",
    "avg_duration_text": "平均待ち時間",
    "median_duration_text": "中央値待ち時間",
    "max_duration": "最大待ち時間",
    "max_duration_text": "最大待ち時間",
    "impact_score": "改善インパクト",
    "impact_share_pct": "改善インパクト比率(%)",
    "case_id": "ケースID",
    "from_time": "開始時刻",
    "to_time": "終了時刻",
    "activity": "Activity",
    "next_activity": "次Activity",
    "transition": "遷移",
    "transition_label": "遷移",
    "duration_text": "所要時間",
    "total_duration": "総所要時間",
    "start_time": "開始時刻",
    "end_time": "終了時刻",
}

DEFAULT_HEADERS = {
    "case_id_column": "case_id",
    "activity_column": "activity",
    "timestamp_column": "start_time",
}

app = FastAPI(title="Process Mining Workbench")
app.mount("/static", StaticFiles(directory=str(BASE_DIR / "static")), name="static")
templates = Jinja2Templates(directory=str(BASE_DIR / "templates"))
RUN_STORE = OrderedDict()


def build_column_selection_payload(headers):
    header_set = set(headers)
    return {
        "headers": headers,
        "default_selection": {
            field_name: (
                default_header if default_header in header_set else ""
            )
            for field_name, default_header in DEFAULT_HEADERS.items()
        },
    }


def validate_selected_columns(case_id_column, activity_column, timestamp_column):
    selected_columns = {
        "Case ID": case_id_column,
        "Activity": activity_column,
        "Timestamp": timestamp_column,
    }

    missing_fields = [
        field_label
        for field_label, column_name in selected_columns.items()
        if not str(column_name or "").strip()
    ]
    if missing_fields:
        raise ValueError(f"Please select: {' / '.join(missing_fields)}")

    normalized_columns = [column_name.strip() for column_name in selected_columns.values()]
    if len(set(normalized_columns)) != len(normalized_columns):
        raise ValueError("Case ID / Activity / Timestamp にはそれぞれ異なる列を選択してください。")


def validate_filter_column_settings(filter_column_settings):
    selected_filter_columns = [
        filter_config["column_name"]
        for filter_config in filter_column_settings.values()
        if filter_config["column_name"]
    ]

    if len(selected_filter_columns) != len(set(selected_filter_columns)):
        raise ValueError("グループ/カテゴリー フィルター①〜③ にはそれぞれ異なる列を選択してください。")


def read_raw_log_dataframe(file_source):
    if hasattr(file_source, "seek"):
        file_source.seek(0)

    try:
        raw_df = pd.read_csv(file_source, dtype=str, keep_default_na=False)
    finally:
        if hasattr(file_source, "seek"):
            file_source.seek(0)

    return raw_df


def resolve_profile_file_source(form):
    uploaded_file = form.get("csv_file")

    if uploaded_file and uploaded_file.filename:
        uploaded_file.file.seek(0)
        return uploaded_file.file, uploaded_file.filename

    return SAMPLE_FILE, SAMPLE_FILE.name


def get_static_version():
    static_dir = BASE_DIR / "static"
    return str(
        max(
            entry.stat().st_mtime_ns
            for entry in static_dir.iterdir()
            if entry.is_file()
        )
    )


def save_run_data(
    source_file_name,
    selected_analysis_keys,
    prepared_df,
    result,
    column_settings,
    base_filter_params,
):
    run_id = uuid4().hex
    RUN_STORE[run_id] = {
        "created_at": datetime.now(timezone.utc).isoformat(),
        "source_file_name": source_file_name,
        "selected_analysis_keys": selected_analysis_keys,
        "prepared_df": prepared_df,
        "result": result,
        "column_settings": column_settings,
        "base_filter_params": base_filter_params,
        "pattern_index_entries": None,
        "pattern_flow_cache": OrderedDict(),
        "variant_cache": {},
        "bottleneck_cache": {},
        "dashboard_cache": {},
        "root_cause_cache": {},
        "impact_cache": {},
        "insights_cache": {},
        "analysis_cache": {},
        "filter_options": None,
    }
    RUN_STORE.move_to_end(run_id)

    while len(RUN_STORE) > MAX_STORED_RUNS:
        RUN_STORE.popitem(last=False)

    return run_id


def get_run_data(run_id):
    run_data = RUN_STORE.get(run_id)

    if not run_data:
        raise HTTPException(status_code=404, detail="Run data was not found.")

    RUN_STORE.move_to_end(run_id)
    return run_data


def get_request_filter_params(request: Request):
    return normalize_filter_params(
        **{
            filter_name: request.query_params.get(filter_name)
            for filter_name in FILTER_PARAM_NAMES
        }
    )


def get_form_filter_params(form):
    return normalize_filter_params(
        **{
            filter_name: form.get(filter_name)
            for filter_name in FILTER_PARAM_NAMES
        }
    )


def get_form_filter_column_settings(form):
    raw_settings = {
        setting_name: form.get(setting_name)
        for setting_name in (*FILTER_COLUMN_NAMES, *FILTER_LABEL_NAMES)
    }
    return normalize_filter_column_settings(**raw_settings)


def get_effective_filter_params(run_data, filter_params=None):
    return merge_filter_params(run_data.get("base_filter_params"), filter_params)


def build_filter_cache_key(filter_params):
    normalized_filters = normalize_filter_params(**(filter_params or {}))
    return tuple(
        normalized_filters.get(filter_name)
        for filter_name in FILTER_PARAM_NAMES
    )


def build_filtered_meta(prepared_df):
    return {
        "case_count": int(prepared_df["case_id"].nunique()) if not prepared_df.empty else 0,
        "event_count": int(len(prepared_df)),
    }


def build_column_settings_payload(column_settings):
    raw_column_settings = column_settings or {}
    filter_slot_names = ("filter_value_1", "filter_value_2", "filter_value_3")

    if any(isinstance(raw_column_settings.get(filter_slot_name), dict) for filter_slot_name in filter_slot_names):
        default_filter_settings = normalize_filter_column_settings()
        normalized_filter_settings = {
            filter_slot_name: {
                "column_name": str((raw_column_settings.get(filter_slot_name) or {}).get("column_name") or "").strip() or None,
                "label": str((raw_column_settings.get(filter_slot_name) or {}).get("label") or "").strip()
                or default_filter_settings[filter_slot_name]["label"],
            }
            for filter_slot_name in filter_slot_names
        }
    else:
        normalized_filter_settings = normalize_filter_column_settings(**raw_column_settings)

    return {
        "case_id_column": str(raw_column_settings.get("case_id_column") or "").strip(),
        "activity_column": str(raw_column_settings.get("activity_column") or "").strip(),
        "timestamp_column": str(raw_column_settings.get("timestamp_column") or "").strip(),
        "filters": [
            {
                "slot": filter_key,
                **normalized_filter_settings[filter_key],
            }
            for filter_key in normalized_filter_settings
        ],
    }


def build_analysis_payload(analysis, row_limit=None, row_offset=0):
    total_row_count = len(analysis["rows"])
    safe_row_offset = max(0, int(row_offset or 0))

    if row_limit is None:
        safe_row_limit = total_row_count
        rows = analysis["rows"][safe_row_offset:]
    else:
        safe_row_limit = max(0, int(row_limit))
        rows = analysis["rows"][safe_row_offset : safe_row_offset + safe_row_limit]

    page_end_row_number = safe_row_offset + len(rows)
    has_previous_page = safe_row_offset > 0
    has_next_page = page_end_row_number < total_row_count
    previous_row_offset = max(0, safe_row_offset - safe_row_limit) if has_previous_page else None
    next_row_offset = page_end_row_number if has_next_page else None

    return {
        "analysis_name": analysis["analysis_name"],
        "sheet_name": analysis["sheet_name"],
        "output_file_name": analysis.get("output_file_name"),
        "row_count": total_row_count,
        "returned_row_count": len(rows),
        "row_offset": safe_row_offset,
        "page_size": safe_row_limit,
        "page_start_row_number": safe_row_offset + 1 if rows else 0,
        "page_end_row_number": page_end_row_number,
        "has_previous_page": has_previous_page,
        "has_next_page": has_next_page,
        "previous_row_offset": previous_row_offset,
        "next_row_offset": next_row_offset,
        "rows": rows,
        "excel_file": analysis["excel_file"],
    }


def build_variant_response_item(variant_item):
    return {
        "variant_id": variant_item["variant_id"],
        "activities": variant_item["activities"],
        "activity_count": variant_item.get("activity_count", len(variant_item["activities"])),
        "pattern": variant_item.get("pattern", ""),
        "count": variant_item["count"],
        "ratio": variant_item["ratio"],
        "avg_case_duration_sec": variant_item.get("avg_case_duration_sec", 0.0),
        "avg_case_duration_text": variant_item.get("avg_case_duration_text", "0s"),
    }


def build_variant_coverage_payload(total_case_count, variant_items):
    covered_case_count = sum(int(variant_item["count"]) for variant_item in variant_items)
    return {
        "displayed_variant_count": len(variant_items),
        "covered_case_count": covered_case_count,
        "total_case_count": int(total_case_count),
        "ratio": round(covered_case_count / total_case_count, 4) if total_case_count else 0.0,
    }


def sanitize_workbook_sheet_name(sheet_name):
    invalid_characters = set('[]:*?/\\')
    normalized_name = "".join("_" if character in invalid_characters else character for character in str(sheet_name or "").strip())
    normalized_name = normalized_name or "Sheet"
    return normalized_name[:31]


def normalize_excel_cell_value(value):
    if value is None:
        return ""

    if isinstance(value, pd.Timestamp):
        return value.isoformat()

    if isinstance(value, (list, tuple, set)):
        return " / ".join(str(item) for item in value)

    if isinstance(value, dict):
        return ", ".join(f"{key}={normalize_excel_cell_value(item_value)}" for key, item_value in value.items())

    try:
        if pd.isna(value):
            return ""
    except TypeError:
        pass

    return value


def autosize_worksheet_columns(worksheet, min_width=10, max_width=48):
    for column_cells in worksheet.columns:
        column_letter = column_cells[0].column_letter
        measured_width = max(
            len(str(normalize_excel_cell_value(cell.value)))
            for cell in column_cells
        ) if column_cells else min_width
        worksheet.column_dimensions[column_letter].width = max(min_width, min(max_width, measured_width + 2))


def append_table_to_worksheet(worksheet, title, rows, headers, start_row=1):
    current_row = start_row
    worksheet.cell(row=current_row, column=1, value=title).font = Font(bold=True)
    current_row += 1

    for column_index, header in enumerate(headers, start=1):
        worksheet.cell(row=current_row, column=column_index, value=header).font = Font(bold=True)
    current_row += 1

    if not rows:
        worksheet.cell(row=current_row, column=1, value="表示できるデータがありません。")
        return current_row + 2

    for row in rows:
        for column_index, header in enumerate(headers, start=1):
            worksheet.cell(
                row=current_row,
                column=column_index,
                value=normalize_excel_cell_value(row.get(header)),
            )
        current_row += 1

    return current_row + 1


def append_key_value_rows(worksheet, title, rows, start_row=1):
    current_row = start_row
    worksheet.cell(row=current_row, column=1, value=title).font = Font(bold=True)
    current_row += 1
    worksheet.cell(row=current_row, column=1, value="項目").font = Font(bold=True)
    worksheet.cell(row=current_row, column=2, value="値").font = Font(bold=True)
    current_row += 1

    for label, value in rows:
        worksheet.cell(row=current_row, column=1, value=label)
        worksheet.cell(row=current_row, column=2, value=normalize_excel_cell_value(value))
        current_row += 1

    return current_row + 1


def build_ranked_rows(rows, rank_key="rank"):
    ranked_rows = []
    for index, row in enumerate(rows, start=1):
        ranked_rows.append({
            rank_key: index,
            **row,
        })
    return ranked_rows


def localize_report_headers(headers):
    return [REPORT_HEADER_LABELS.get(header, header) for header in headers]


def localize_report_rows(rows, headers):
    localized_headers = localize_report_headers(headers)
    localized_rows = []

    for row in rows:
        localized_rows.append(
            {
                localized_header: row.get(header)
                for header, localized_header in zip(headers, localized_headers)
            }
        )

    return localized_rows, localized_headers


def build_filter_summary_text(filter_params, column_settings):
    normalized_filters = normalize_filter_params(**(filter_params or {}))
    column_payload = build_column_settings_payload(column_settings)
    filter_label_map = {
        filter_item["slot"]: filter_item["label"]
        for filter_item in column_payload.get("filters", [])
    }
    summary_items = []

    if normalized_filters.get("date_from"):
        summary_items.append(f"開始日: {normalized_filters['date_from']}")
    if normalized_filters.get("date_to"):
        summary_items.append(f"終了日: {normalized_filters['date_to']}")

    for filter_slot in ("filter_value_1", "filter_value_2", "filter_value_3"):
        if normalized_filters.get(filter_slot):
            summary_items.append(
                f"{filter_label_map.get(filter_slot, filter_slot)}: {normalized_filters[filter_slot]}"
            )

    activity_values = normalized_filters.get("activity_values")
    if activity_values:
        activity_label = "Activity 含む" if normalized_filters.get("activity_mode") != "exclude" else "Activity 除外"
        summary_items.append(f"{activity_label}: {activity_values}")

    return " / ".join(summary_items) if summary_items else "未適用"


def parse_transition_selection(selected_transition_key):
    normalized_key = str(selected_transition_key or "").strip()
    if "__TO__" not in normalized_key:
        return "", ""
    from_activity, to_activity = normalized_key.split("__TO__", 1)
    return from_activity.strip(), to_activity.strip()


def build_detail_export_workbook_bytes(
    run_id,
    run_data,
    analysis_key,
    filter_params,
    variant_id=None,
    selected_activity="",
    selected_transition_key="",
    case_id="",
    drilldown_limit=20,
):
    analysis_definitions = get_available_analysis_definitions()
    analysis_name = analysis_definitions.get(analysis_key, {}).get("config", {}).get("analysis_name", analysis_key)
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = sanitize_workbook_sheet_name(REPORT_SHEET_NAMES["summary"])

    filtered_df = filter_prepared_df(
        run_data["prepared_df"],
        filter_params,
        filter_column_settings=run_data.get("column_settings"),
    )
    filtered_meta = build_filtered_meta(filtered_df)
    frequency_analysis = get_analysis_data(run_data, "frequency", filter_params=filter_params)
    pattern_analysis = get_analysis_data(run_data, "pattern", filter_params=filter_params)

    from_activity, to_activity = parse_transition_selection(selected_transition_key)
    selected_transition_label = f"{from_activity} → {to_activity}" if from_activity and to_activity else str(selected_transition_key or "").strip()
    summary_rows = [
        (REPORT_HEADER_LABELS["run_id"], run_id),
        (REPORT_HEADER_LABELS["analysis_key"], analysis_key),
        (REPORT_HEADER_LABELS["analysis_name"], analysis_name),
        (REPORT_HEADER_LABELS["source_file_name"], run_data["source_file_name"]),
        (REPORT_HEADER_LABELS["analysis_executed_at"], run_data.get("created_at", "")),
        (REPORT_HEADER_LABELS["exported_at"], datetime.now(timezone.utc).isoformat()),
        (REPORT_HEADER_LABELS["case_count"], filtered_meta["case_count"]),
        (REPORT_HEADER_LABELS["event_count"], filtered_meta["event_count"]),
        (REPORT_HEADER_LABELS["applied_filters"], build_filter_summary_text(filter_params, run_data.get("column_settings"))),
        (REPORT_HEADER_LABELS["selected_variant"], f"Variant #{variant_id}" if variant_id else "未選択"),
        (REPORT_HEADER_LABELS["selected_activity"], selected_activity or "未選択"),
        (REPORT_HEADER_LABELS["selected_transition"], selected_transition_label or "未選択"),
        (REPORT_HEADER_LABELS["selected_case_id"], case_id or "未選択"),
    ]
    append_key_value_rows(summary_sheet, REPORT_SHEET_NAMES["summary"], summary_rows)

    frequency_sheet = workbook.create_sheet(title=sanitize_workbook_sheet_name(REPORT_SHEET_NAMES["frequency"]))
    frequency_rows = build_ranked_rows(frequency_analysis["rows"], rank_key=REPORT_HEADER_LABELS["rank"])
    frequency_headers = list(frequency_rows[0].keys()) if frequency_rows else [REPORT_HEADER_LABELS["rank"]]
    append_table_to_worksheet(
        frequency_sheet,
        REPORT_SHEET_NAMES["frequency"],
        frequency_rows,
        frequency_headers,
    )

    pattern_sheet = workbook.create_sheet(title=sanitize_workbook_sheet_name(REPORT_SHEET_NAMES["pattern"]))
    pattern_rows = build_ranked_rows(pattern_analysis["rows"], rank_key=REPORT_HEADER_LABELS["rank"])
    pattern_headers = list(pattern_rows[0].keys()) if pattern_rows else [REPORT_HEADER_LABELS["rank"]]
    append_table_to_worksheet(
        pattern_sheet,
        REPORT_SHEET_NAMES["pattern"],
        pattern_rows,
        pattern_headers,
    )

    variant_sheet = workbook.create_sheet(title=sanitize_workbook_sheet_name(REPORT_SHEET_NAMES["variant"]))
    variant_rows, variant_headers = localize_report_rows(
        build_ranked_rows([
            {
                "variant_id": variant_item["variant_id"],
                "count": variant_item["count"],
                "ratio": variant_item["ratio"],
                "avg_case_duration": variant_item.get("avg_case_duration_text", "0s"),
                "pattern": variant_item.get("pattern", ""),
            }
            for variant_item in get_variant_items(run_data, filter_params=filter_params)
        ]),
        ["rank", "variant_id", "count", "ratio", "avg_case_duration", "pattern"],
    )
    append_table_to_worksheet(
        variant_sheet,
        REPORT_SHEET_NAMES["variant"],
        variant_rows,
        variant_headers,
    )

    bottleneck_sheet = workbook.create_sheet(title=sanitize_workbook_sheet_name(REPORT_SHEET_NAMES["bottleneck"]))
    bottleneck_summary = get_bottleneck_summary(
        run_data,
        variant_id=variant_id,
        filter_params=filter_params,
    )
    activity_bottleneck_rows, activity_bottleneck_headers = localize_report_rows(
        bottleneck_summary["activity_bottlenecks"],
        ["rank", "activity", "count", "case_count", "avg_duration_text", "median_duration_text", "max_duration_text"],
    )
    next_row = append_table_to_worksheet(
        bottleneck_sheet,
        "Activityボトルネック",
        activity_bottleneck_rows,
        activity_bottleneck_headers,
    )
    transition_bottleneck_rows, transition_bottleneck_headers = localize_report_rows(
        bottleneck_summary["transition_bottlenecks"],
        ["rank", "transition_label", "count", "case_count", "avg_duration_text", "median_duration_text", "max_duration_text"],
    )
    append_table_to_worksheet(
        bottleneck_sheet,
        "Transitionボトルネック",
        transition_bottleneck_rows,
        transition_bottleneck_headers,
        start_row=next_row,
    )

    impact_sheet = workbook.create_sheet(title=sanitize_workbook_sheet_name(REPORT_SHEET_NAMES["impact"]))
    impact_summary = get_impact_summary(
        run_data,
        filter_params=filter_params,
    )
    impact_rows, impact_headers = localize_report_rows(
        [
            {
                "rank": impact_row["rank"],
                "transition": impact_row["transition_label"],
                "case_count": impact_row["case_count"],
                "avg_duration": impact_row["avg_duration_text"],
                "max_duration": impact_row["max_duration_text"],
                "impact_score": impact_row["impact_score"],
                "impact_share_pct": impact_row["impact_share_pct"],
            }
            for impact_row in impact_summary["rows"]
        ],
        ["rank", "transition", "case_count", "avg_duration", "max_duration", "impact_score", "impact_share_pct"],
    )
    append_table_to_worksheet(
        impact_sheet,
        REPORT_SHEET_NAMES["impact"],
        impact_rows,
        impact_headers,
    )

    selected_activity_name = str(selected_activity or "").strip()
    drilldown_df = get_filtered_prepared_df(
        run_data,
        variant_id=variant_id,
        filter_params=filter_params,
    )
    drilldown_rows = []
    drilldown_title = REPORT_SHEET_NAMES["drilldown"]
    if from_activity and to_activity:
        drilldown_title = f"遷移ドリルダウン: {from_activity} → {to_activity}"
        drilldown_rows = create_transition_case_drilldown(
            drilldown_df,
            from_activity=from_activity,
            to_activity=to_activity,
            limit=max(0, int(drilldown_limit)),
        )
    elif selected_activity_name:
        drilldown_title = f"Activityドリルダウン: {selected_activity_name}"
        drilldown_rows = create_activity_case_drilldown(
            drilldown_df,
            activity=selected_activity_name,
            limit=max(0, int(drilldown_limit)),
        )
    if drilldown_rows:
        drilldown_sheet = workbook.create_sheet(title=sanitize_workbook_sheet_name(REPORT_SHEET_NAMES["drilldown"]))
        drilldown_rows, drilldown_headers = localize_report_rows(
            drilldown_rows,
            ["case_id", "activity", "next_activity", "duration_text", "from_time", "to_time"],
        )
        append_table_to_worksheet(
            drilldown_sheet,
            drilldown_title,
            drilldown_rows,
            drilldown_headers,
        )

    normalized_case_id = str(case_id or "").strip()
    if normalized_case_id:
        case_trace = create_case_trace_details(run_data["prepared_df"], normalized_case_id)
        if case_trace.get("found"):
            case_trace_sheet = workbook.create_sheet(title=sanitize_workbook_sheet_name(REPORT_SHEET_NAMES["case_trace"]))
            next_row = append_key_value_rows(
                case_trace_sheet,
                "ケース概要",
                [
                    (REPORT_HEADER_LABELS["case_id"], case_trace["case_id"]),
                    (REPORT_HEADER_LABELS["event_count"], case_trace["summary"]["event_count"]),
                    (REPORT_HEADER_LABELS["total_duration"], case_trace["summary"]["total_duration_text"]),
                    (REPORT_HEADER_LABELS["start_time"], case_trace["summary"]["start_time"]),
                    (REPORT_HEADER_LABELS["end_time"], case_trace["summary"]["end_time"]),
                ],
            )
            case_trace_event_rows, case_trace_event_headers = localize_report_rows(
                case_trace["events"],
                ["case_id", "activity", "next_activity", "start_time", "end_time", "duration_text"],
            )
            append_table_to_worksheet(
                case_trace_sheet,
                "通過イベント",
                case_trace_event_rows,
                case_trace_event_headers,
                start_row=next_row,
            )

    for worksheet in workbook.worksheets:
        autosize_worksheet_columns(worksheet)

    output_buffer = BytesIO()
    workbook.save(output_buffer)
    return output_buffer.getvalue()


def get_filter_options_payload(run_data):
    filter_options = run_data.get("filter_options")
    if filter_options is None:
        filter_options = get_filter_options(
            run_data["prepared_df"],
            filter_column_settings=run_data.get("column_settings"),
        )
        run_data["filter_options"] = filter_options

    return filter_options


def get_variant_items(run_data, filter_params=None):
    cache_key = build_filter_cache_key(filter_params)
    variant_cache = run_data.setdefault("variant_cache", {})

    if cache_key not in variant_cache:
        filtered_df = filter_prepared_df(
            run_data["prepared_df"],
            filter_params,
            filter_column_settings=run_data.get("column_settings"),
        )
        variant_cache[cache_key] = create_variant_summary(filtered_df, limit=None)

    return variant_cache[cache_key]


def get_variant_item(run_data, variant_id, filter_params=None):
    safe_variant_id = int(variant_id)

    for variant_item in get_variant_items(run_data, filter_params=filter_params):
        if variant_item["variant_id"] == safe_variant_id:
            return variant_item

    raise HTTPException(status_code=404, detail="Variant was not found.")


def get_pattern_summary_row(run_data, pattern_index):
    pattern_analysis = run_data["result"]["analyses"].get("pattern")

    if not pattern_analysis:
        raise HTTPException(status_code=400, detail="Pattern analysis is not available.")

    pattern_rows = pattern_analysis["rows"]
    safe_pattern_index = int(pattern_index)
    if safe_pattern_index < 0 or safe_pattern_index >= len(pattern_rows):
        raise HTTPException(status_code=404, detail="Pattern index was not found.")

    pattern_index_entries = run_data.get("pattern_index_entries")
    if pattern_index_entries is None:
        pattern_index_entries = create_pattern_index_entries(run_data["prepared_df"])
        run_data["pattern_index_entries"] = pattern_index_entries

    if safe_pattern_index >= len(pattern_index_entries):
        raise HTTPException(status_code=404, detail="Pattern index was not found.")

    summary_row = pattern_rows[safe_pattern_index]
    pattern_entry = pattern_index_entries[safe_pattern_index]
    pattern = str(pattern_entry.get("pattern") or "").strip()

    if not pattern:
        raise HTTPException(status_code=500, detail="Pattern text could not be resolved.")

    return pattern_analysis, summary_row, pattern_entry, pattern


def get_filtered_prepared_df(run_data, variant_id=None, pattern_index=None, filter_params=None):
    prepared_df = filter_prepared_df(
        run_data["prepared_df"],
        filter_params,
        filter_column_settings=run_data.get("column_settings"),
    )

    if pattern_index is not None:
        _, _, _, pattern = get_pattern_summary_row(run_data, pattern_index)
        prepared_df = filter_prepared_df_by_pattern(prepared_df, pattern)

    if variant_id is not None:
        variant_item = get_variant_item(run_data, variant_id, filter_params=filter_params)
        prepared_df = filter_prepared_df_by_pattern(prepared_df, variant_item["pattern"])

    return prepared_df


def get_analysis_data(run_data, analysis_key, filter_params=None):
    normalized_filter_key = build_filter_cache_key(filter_params)
    analysis_cache = run_data.setdefault("analysis_cache", {})

    if all(filter_value is None for filter_value in normalized_filter_key):
        analysis = run_data["result"]["analyses"].get(analysis_key)
        if analysis:
            return analysis

        cache_key = ("analysis", analysis_key, normalized_filter_key)
        if cache_key not in analysis_cache:
            analysis_cache[cache_key] = create_analysis_records(run_data["prepared_df"], analysis_key)
        return analysis_cache[cache_key]

    cache_key = ("analysis", analysis_key, normalized_filter_key)

    if cache_key not in analysis_cache:
        filtered_df = filter_prepared_df(
            run_data["prepared_df"],
            filter_params,
            filter_column_settings=run_data.get("column_settings"),
        )
        analysis_cache[cache_key] = create_analysis_records(filtered_df, analysis_key)

    return analysis_cache[cache_key]


def get_bottleneck_summary(run_data, variant_id=None, pattern_index=None, filter_params=None):
    cache_key = (
        "bottleneck",
        None if variant_id is None else int(variant_id),
        None if pattern_index is None else int(pattern_index),
        build_filter_cache_key(filter_params),
    )
    cache = run_data.setdefault("bottleneck_cache", {})

    if cache_key not in cache:
        filtered_df = get_filtered_prepared_df(
            run_data,
            variant_id=variant_id,
            pattern_index=pattern_index,
            filter_params=filter_params,
        )
        cache[cache_key] = create_bottleneck_summary(filtered_df, limit=None)

    return cache[cache_key]


def get_dashboard_summary(run_data, filter_params=None, prepared_df=None):
    cache_key = build_filter_cache_key(filter_params)
    cache = run_data.setdefault("dashboard_cache", {})

    if cache_key not in cache:
        filtered_df = prepared_df
        if filtered_df is None:
            filtered_df = filter_prepared_df(
                run_data["prepared_df"],
                filter_params,
                filter_column_settings=run_data.get("column_settings"),
            )

        cache[cache_key] = create_dashboard_summary(
            filtered_df,
            variant_items=get_variant_items(run_data, filter_params=filter_params)[:10],
            bottleneck_summary=get_bottleneck_summary(run_data, filter_params=filter_params),
            coverage_limit=10,
        )

    return cache[cache_key]


def get_root_cause_summary(run_data, filter_params=None, prepared_df=None):
    cache_key = build_filter_cache_key(filter_params)
    cache = run_data.setdefault("root_cause_cache", {})

    if cache_key not in cache:
        filtered_df = prepared_df
        if filtered_df is None:
            filtered_df = filter_prepared_df(
                run_data["prepared_df"],
                filter_params,
                filter_column_settings=run_data.get("column_settings"),
            )

        cache[cache_key] = create_root_cause_summary(
            filtered_df,
            filter_column_settings=run_data.get("column_settings"),
            limit=10,
        )

    return cache[cache_key]


def get_impact_summary(run_data, filter_params=None, prepared_df=None):
    cache_key = build_filter_cache_key(filter_params)
    cache = run_data.setdefault("impact_cache", {})

    if cache_key not in cache:
        filtered_df = prepared_df
        if filtered_df is None:
            filtered_df = filter_prepared_df(
                run_data["prepared_df"],
                filter_params,
                filter_column_settings=run_data.get("column_settings"),
            )

        cache[cache_key] = create_impact_summary(
            filtered_df,
            limit=None,
        )

    return cache[cache_key]


def get_rule_based_insights_summary(
    run_data,
    analysis_key,
    analysis_rows=None,
    filter_params=None,
    prepared_df=None,
    dashboard_summary=None,
    impact_summary=None,
):
    cache_key = (str(analysis_key or "").strip().lower(), build_filter_cache_key(filter_params))
    cache = run_data.setdefault("insights_cache", {})

    if cache_key not in cache:
        filtered_df = prepared_df
        if filtered_df is None:
            filtered_df = filter_prepared_df(
                run_data["prepared_df"],
                filter_params,
                filter_column_settings=run_data.get("column_settings"),
            )

        cache[cache_key] = create_rule_based_insights(
            filtered_df,
            analysis_key=analysis_key,
            analysis_rows=analysis_rows,
            dashboard_summary=dashboard_summary or get_dashboard_summary(
                run_data,
                filter_params=filter_params,
                prepared_df=filtered_df,
            ),
            bottleneck_summary=get_bottleneck_summary(
                run_data,
                filter_params=filter_params,
            ),
            impact_summary=impact_summary or get_impact_summary(
                run_data,
                filter_params=filter_params,
                prepared_df=filtered_df,
            ),
            max_items=5,
        )

    return cache[cache_key]


def get_pattern_flow_snapshot(
    run_data,
    pattern_percent,
    pattern_count,
    activity_percent,
    connection_percent,
    variant_id=None,
    filter_params=None,
):
    filter_cache_key = build_filter_cache_key(filter_params)
    if variant_id is None:
        cache_key = (
            "pattern",
            int(pattern_percent),
            None if pattern_count is None else int(pattern_count),
            int(activity_percent),
            int(connection_percent),
            filter_cache_key,
        )
    else:
        cache_key = (
            "variant",
            int(variant_id),
            int(activity_percent),
            int(connection_percent),
            filter_cache_key,
        )
    cache = run_data.setdefault("pattern_flow_cache", OrderedDict())

    cached_snapshot = cache.get(cache_key)
    if cached_snapshot is not None:
        cache.move_to_end(cache_key)
        return cached_snapshot

    analyses = run_data["result"]["analyses"]
    if variant_id is None:
        pattern_analysis = get_analysis_data(run_data, "pattern", filter_params=filter_params)
        frequency_analysis = get_analysis_data(run_data, "frequency", filter_params=filter_params)
        snapshot = create_pattern_flow_snapshot(
            pattern_rows=pattern_analysis["rows"],
            frequency_rows=frequency_analysis.get("rows", []),
            pattern_percent=pattern_percent,
            pattern_count=pattern_count,
            activity_percent=activity_percent,
            connection_percent=connection_percent,
            pattern_cap=PROCESS_FLOW_PATTERN_CAP,
        )
    else:
        filtered_df = filter_prepared_df(
            run_data["prepared_df"],
            filter_params,
            filter_column_settings=run_data.get("column_settings"),
        )
        variant_item = get_variant_item(run_data, variant_id, filter_params=filter_params)
        snapshot = create_variant_flow_snapshot(
            prepared_df=filtered_df,
            variant_pattern=variant_item["pattern"],
            activity_percent=activity_percent,
            connection_percent=connection_percent,
        )
        snapshot["selected_variant"] = build_variant_response_item(variant_item)

    cache[cache_key] = snapshot
    cache.move_to_end(cache_key)
    while len(cache) > MAX_PATTERN_FLOW_CACHE:
        cache.popitem(last=False)

    return snapshot


def build_preview_response(run_id, source_file_name, selected_analysis_keys, result, run_data):
    return {
        "run_id": run_id,
        "source_file_name": source_file_name,
        "selected_analysis_keys": selected_analysis_keys,
        "case_count": result["case_count"],
        "event_count": result["event_count"],
        "applied_filters": run_data.get("base_filter_params"),
        "column_settings": build_column_settings_payload(run_data.get("column_settings")),
        "filter_options": get_filter_options_payload(run_data),
        "analyses": {
            analysis_key: build_analysis_payload(analysis, PREVIEW_ROW_COUNT)
            for analysis_key, analysis in result["analyses"].items()
        },
    }


def build_log_profile_payload(
    raw_df,
    source_file_name,
    case_id_column="",
    activity_column="",
    timestamp_column="",
    filter_column_settings=None,
    include_diagnostics=False,
):
    headers = [str(column_name) for column_name in raw_df.columns.tolist()]
    selection_payload = build_column_selection_payload(headers)
    resolved_case_id_column = case_id_column or selection_payload["default_selection"]["case_id_column"]
    resolved_activity_column = activity_column or selection_payload["default_selection"]["activity_column"]
    resolved_timestamp_column = timestamp_column or selection_payload["default_selection"]["timestamp_column"]
    normalized_filter_column_settings = normalize_filter_column_settings(**(filter_column_settings or {}))

    return {
        "source_file_name": source_file_name,
        **selection_payload,
        "column_settings": build_column_settings_payload(
            {
                "case_id_column": resolved_case_id_column,
                "activity_column": resolved_activity_column,
                "timestamp_column": resolved_timestamp_column,
                **normalized_filter_column_settings,
            }
        ),
        "filter_options": get_filter_options(
            raw_df,
            filter_column_settings=normalized_filter_column_settings,
        ),
        "diagnostics": (
            create_log_diagnostics(
                raw_df,
                case_id_column=resolved_case_id_column,
                activity_column=resolved_activity_column,
                timestamp_column=resolved_timestamp_column,
                filter_column_settings=normalized_filter_column_settings,
            )
            if include_diagnostics
            else None
        ),
    }


def get_analysis_options():
    analysis_definitions = get_available_analysis_definitions()
    analysis_options = []

    for analysis_key in DEFAULT_ANALYSIS_KEYS:
        analysis_options.append(
            {
                "key": analysis_key,
                "label": analysis_definitions[analysis_key]["config"]["analysis_name"],
            }
        )

    return analysis_options


@app.get("/")
def index(request: Request):
    sample_profile_payload = build_log_profile_payload(
        raw_df=read_raw_log_dataframe(SAMPLE_FILE),
        source_file_name=SAMPLE_FILE.name,
        include_diagnostics=False,
    )

    return templates.TemplateResponse(
        request,
        "index.html",
        context={
            "analysis_options": get_analysis_options(),
            "default_headers": DEFAULT_HEADERS,
            "sample_profile_payload": sample_profile_payload,
            "sample_file_name": SAMPLE_FILE.name,
            "static_version": get_static_version(),
        },
    )


@app.get("/analysis/patterns/{pattern_index}")
def pattern_detail_page(request: Request, pattern_index: int):
    return templates.TemplateResponse(
        request,
        "pattern_detail.html",
        context={
            "pattern_index": pattern_index,
            "static_version": get_static_version(),
        },
    )


@app.get("/analysis/{analysis_key}")
def analysis_detail(request: Request, analysis_key):
    analysis_definitions = get_available_analysis_definitions()

    if analysis_key not in analysis_definitions:
        raise HTTPException(status_code=404, detail="Analysis key was not found.")

    return templates.TemplateResponse(
        request,
        "analysis_detail.html",
        context={
            "analysis_key": analysis_key,
            "analysis_name": analysis_definitions[analysis_key]["config"]["analysis_name"],
            "static_version": get_static_version(),
        },
    )


@app.get("/api/runs/{run_id}/patterns/{pattern_index}")
def pattern_detail_api(run_id: str, pattern_index: int):
    run_data = get_run_data(run_id)
    pattern_analysis, summary_row, _, pattern = get_pattern_summary_row(run_data, pattern_index)

    detail = create_pattern_bottleneck_details(run_data["prepared_df"], pattern)
    return JSONResponse(
        content={
            "run_id": run_id,
            "pattern_index": pattern_index,
            "source_file_name": run_data["source_file_name"],
            "analysis_name": pattern_analysis["analysis_name"],
            "summary_row": summary_row,
            **detail,
        }
    )


@app.get("/api/runs/{run_id}/analyses/{analysis_key}")
def analysis_detail_api(
    request: Request,
    run_id: str,
    analysis_key: str,
    row_limit: int | None = None,
    row_offset: int = 0,
):
    run_data = get_run_data(run_id)
    filter_params = get_effective_filter_params(run_data, get_request_filter_params(request))
    analysis = get_analysis_data(run_data, analysis_key, filter_params=filter_params)
    filtered_df = filter_prepared_df(
        run_data["prepared_df"],
        filter_params,
        filter_column_settings=run_data.get("column_settings"),
    )
    filtered_meta = build_filtered_meta(filtered_df)

    response_analyses = {
        analysis_key: build_analysis_payload(
            analysis,
            row_limit=row_limit,
            row_offset=row_offset,
        )
    }
    dashboard_summary = get_dashboard_summary(
        run_data,
        filter_params=filter_params,
        prepared_df=filtered_df,
    )
    impact_summary = get_impact_summary(
        run_data,
        filter_params=filter_params,
        prepared_df=filtered_df,
    )
    root_cause_summary = get_root_cause_summary(
        run_data,
        filter_params=filter_params,
        prepared_df=filtered_df,
    )

    return JSONResponse(
        content={
            "run_id": run_id,
            "source_file_name": run_data["source_file_name"],
            "selected_analysis_keys": run_data["selected_analysis_keys"],
            "case_count": filtered_meta["case_count"],
            "event_count": filtered_meta["event_count"],
            "dashboard": dashboard_summary,
            "impact": impact_summary,
            "insights": get_rule_based_insights_summary(
                run_data,
                analysis_key=analysis_key,
                analysis_rows=analysis.get("rows"),
                filter_params=filter_params,
                prepared_df=filtered_df,
                dashboard_summary=dashboard_summary,
                impact_summary=impact_summary,
            ),
            "root_cause": root_cause_summary,
            "applied_filters": filter_params,
            "column_settings": build_column_settings_payload(run_data.get("column_settings")),
            "analyses": response_analyses,
        }
    )


@app.get("/api/runs/{run_id}/excel-files/{analysis_key}")
def analysis_excel_file_api(run_id: str, analysis_key: str):
    run_data = get_run_data(run_id)
    analyses = run_data["result"]["analyses"]
    analysis = analyses.get(analysis_key)

    if not analysis:
        raise HTTPException(status_code=404, detail="Analysis data was not found.")

    excel_df = pd.DataFrame(analysis["rows"])
    excel_bytes = build_excel_bytes(excel_df, analysis["sheet_name"])
    output_file_name = analysis.get("output_file_name") or f"{analysis['analysis_name']}.xlsx"

    return Response(
        content=excel_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{quote(output_file_name)}",
        },
    )


@app.get("/api/runs/{run_id}/excel-archive")
def analysis_excel_archive_api(run_id: str):
    run_data = get_run_data(run_id)
    analyses = run_data["result"]["analyses"]

    archive_buffer = BytesIO()
    with ZipFile(archive_buffer, mode="w", compression=ZIP_DEFLATED) as archive_file:
        for analysis in analyses.values():
            excel_df = pd.DataFrame(analysis["rows"])
            output_file_name = analysis.get("output_file_name") or f"{analysis['analysis_name']}.xlsx"
            excel_bytes = build_excel_bytes(excel_df, analysis["sheet_name"])
            archive_file.writestr(output_file_name, excel_bytes)

    archive_file_name = f"{Path(run_data['source_file_name']).stem}_analysis_excels.zip"

    return Response(
        content=archive_buffer.getvalue(),
        media_type="application/zip",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{quote(archive_file_name)}",
        },
    )


@app.get("/api/runs/{run_id}/detail-excel")
@app.get("/api/runs/{run_id}/report-excel")
def detail_excel_export_api(
    request: Request,
    run_id: str,
    analysis_key: str,
    variant_id: int | None = None,
    selected_activity: str = "",
    selected_transition_key: str = "",
    case_id: str = "",
    drilldown_limit: int = 20,
):
    run_data = get_run_data(run_id)
    filter_params = get_effective_filter_params(run_data, get_request_filter_params(request))
    excel_bytes = build_detail_export_workbook_bytes(
        run_id=run_id,
        run_data=run_data,
        analysis_key=analysis_key,
        filter_params=filter_params,
        variant_id=variant_id,
        selected_activity=selected_activity,
        selected_transition_key=selected_transition_key,
        case_id=case_id,
        drilldown_limit=drilldown_limit,
    )
    output_file_name = f"{Path(run_data['source_file_name']).stem}_{analysis_key}_detail.xlsx"

    return Response(
        content=excel_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{quote(output_file_name)}",
        },
    )


@app.get("/api/runs/{run_id}/filter-options")
def filter_options_api(run_id: str):
    run_data = get_run_data(run_id)

    return JSONResponse(
        content={
            "run_id": run_id,
            "options": get_filter_options_payload(run_data),
            "applied_filters": run_data.get("base_filter_params"),
            "column_settings": build_column_settings_payload(run_data.get("column_settings")),
        }
    )


@app.get("/api/runs/{run_id}/pattern-flow")
def pattern_flow_api(
    request: Request,
    run_id: str,
    pattern_percent: int = 10,
    pattern_count: int | None = None,
    activity_percent: int = 40,
    connection_percent: int = 30,
    variant_id: int | None = None,
):
    run_data = get_run_data(run_id)
    filter_params = get_effective_filter_params(run_data, get_request_filter_params(request))
    pattern_analysis = get_analysis_data(run_data, "pattern", filter_params=filter_params)

    if not pattern_analysis and variant_id is None:
        raise HTTPException(status_code=400, detail="Pattern analysis is not available.")

    snapshot = get_pattern_flow_snapshot(
        run_data=run_data,
        pattern_percent=pattern_percent,
        pattern_count=pattern_count,
        activity_percent=activity_percent,
        connection_percent=connection_percent,
        variant_id=variant_id,
        filter_params=filter_params,
    )
    filtered_meta = build_filtered_meta(
        filter_prepared_df(
            run_data["prepared_df"],
            filter_params,
            filter_column_settings=run_data.get("column_settings"),
        )
    )

    return JSONResponse(
        content={
            "run_id": run_id,
            "filtered_case_count": filtered_meta["case_count"],
            "filtered_event_count": filtered_meta["event_count"],
            "applied_filters": filter_params,
            **snapshot,
        }
    )


@app.get("/api/runs/{run_id}/variants")
def variant_list_api(request: Request, run_id: str, limit: int = 10):
    run_data = get_run_data(run_id)
    filter_params = get_effective_filter_params(run_data, get_request_filter_params(request))
    safe_limit = max(0, int(limit))
    filtered_df = filter_prepared_df(
        run_data["prepared_df"],
        filter_params,
        filter_column_settings=run_data.get("column_settings"),
    )
    all_variant_items = get_variant_items(run_data, filter_params=filter_params)
    variant_items = all_variant_items if safe_limit == 0 else all_variant_items[:safe_limit]

    return JSONResponse(
        content={
            "run_id": run_id,
            "variants": [
                build_variant_response_item(variant_item)
                for variant_item in variant_items
            ],
            "coverage": build_variant_coverage_payload(
                total_case_count=int(filtered_df["case_id"].nunique()) if not filtered_df.empty else 0,
                variant_items=variant_items,
            ),
            "filtered_case_count": int(filtered_df["case_id"].nunique()) if not filtered_df.empty else 0,
            "filtered_event_count": int(len(filtered_df)),
            "applied_filters": filter_params,
        }
    )


@app.get("/api/runs/{run_id}/bottlenecks")
def bottleneck_list_api(
    request: Request,
    run_id: str,
    limit: int = 5,
    variant_id: int | None = None,
    pattern_index: int | None = None,
):
    run_data = get_run_data(run_id)
    filter_params = get_effective_filter_params(run_data, get_request_filter_params(request))
    safe_limit = max(0, int(limit))
    bottleneck_summary = get_bottleneck_summary(
        run_data,
        variant_id=variant_id,
        pattern_index=pattern_index,
        filter_params=filter_params,
    )
    filtered_df = get_filtered_prepared_df(
        run_data,
        variant_id=variant_id,
        pattern_index=pattern_index,
        filter_params=filter_params,
    )

    return JSONResponse(
        content={
            "run_id": run_id,
            "limit": safe_limit,
            "variant_id": variant_id,
            "pattern_index": pattern_index,
            "filtered_case_count": int(filtered_df["case_id"].nunique()) if not filtered_df.empty else 0,
            "filtered_event_count": int(len(filtered_df)),
            "applied_filters": filter_params,
            "activity_bottlenecks": bottleneck_summary["activity_bottlenecks"][:safe_limit],
            "transition_bottlenecks": bottleneck_summary["transition_bottlenecks"][:safe_limit],
            "activity_heatmap": bottleneck_summary["activity_heatmap"],
            "transition_heatmap": bottleneck_summary["transition_heatmap"],
        }
    )


@app.get("/api/runs/{run_id}/transition-cases")
def transition_case_drilldown_api(
    request: Request,
    run_id: str,
    from_activity: str,
    to_activity: str,
    limit: int = 20,
    variant_id: int | None = None,
    pattern_index: int | None = None,
):
    run_data = get_run_data(run_id)
    filter_params = get_effective_filter_params(run_data, get_request_filter_params(request))
    filtered_df = get_filtered_prepared_df(
        run_data,
        variant_id=variant_id,
        pattern_index=pattern_index,
        filter_params=filter_params,
    )
    safe_limit = max(0, int(limit))
    case_rows = create_transition_case_drilldown(
        filtered_df,
        from_activity=from_activity,
        to_activity=to_activity,
        limit=safe_limit,
    )

    return JSONResponse(
        content={
            "run_id": run_id,
            "variant_id": variant_id,
            "pattern_index": pattern_index,
            "from_activity": from_activity,
            "to_activity": to_activity,
            "transition_key": f"{from_activity}__TO__{to_activity}",
            "transition_label": f"{from_activity} → {to_activity}",
            "limit": safe_limit,
            "returned_case_count": len(case_rows),
            "applied_filters": filter_params,
            "cases": case_rows,
        }
    )


@app.get("/api/runs/{run_id}/activity-cases")
def activity_case_drilldown_api(
    request: Request,
    run_id: str,
    activity: str,
    limit: int = 20,
    variant_id: int | None = None,
    pattern_index: int | None = None,
):
    run_data = get_run_data(run_id)
    filter_params = get_effective_filter_params(run_data, get_request_filter_params(request))
    filtered_df = get_filtered_prepared_df(
        run_data,
        variant_id=variant_id,
        pattern_index=pattern_index,
        filter_params=filter_params,
    )
    safe_limit = max(0, int(limit))
    case_rows = create_activity_case_drilldown(
        filtered_df,
        activity=activity,
        limit=safe_limit,
    )

    return JSONResponse(
        content={
            "run_id": run_id,
            "variant_id": variant_id,
            "pattern_index": pattern_index,
            "activity": activity,
            "limit": safe_limit,
            "returned_case_count": len(case_rows),
            "applied_filters": filter_params,
            "cases": case_rows,
        }
    )


@app.get("/api/runs/{run_id}/cases/{case_id:path}")
def case_trace_api(run_id: str, case_id: str):
    run_data = get_run_data(run_id)
    normalized_case_id = str(case_id or "").strip()

    if not normalized_case_id:
        return JSONResponse(
            status_code=400,
            content={
                "run_id": run_id,
                "case_id": "",
                "found": False,
                "summary": None,
                "events": [],
                "error": "Case ID is required.",
            },
        )

    case_trace = create_case_trace_details(run_data["prepared_df"], normalized_case_id)
    return JSONResponse(
        content={
            "run_id": run_id,
            **case_trace,
        }
    )


@app.post("/api/csv-headers")
async def csv_headers(request: Request):
    form = await request.form()
    raw_case_id_column = form.get("case_id_column")
    raw_activity_column = form.get("activity_column")
    raw_timestamp_column = form.get("timestamp_column")
    filter_column_settings = get_form_filter_column_settings(form)
    file_source, source_file_name = resolve_profile_file_source(form)

    try:
        raw_df = read_raw_log_dataframe(file_source)
    except ValueError as exc:
        return JSONResponse(status_code=400, content={"error": str(exc)})
    except Exception as exc:
        return JSONResponse(
            status_code=400,
            content={
                "error": "CSV headers could not be read. Please check the file encoding and header row.",
                "detail": str(exc),
            },
        )

    return JSONResponse(
        content=build_log_profile_payload(
            raw_df=raw_df,
            source_file_name=source_file_name,
            case_id_column=str(raw_case_id_column or "").strip(),
            activity_column=str(raw_activity_column or "").strip(),
            timestamp_column=str(raw_timestamp_column or "").strip(),
            filter_column_settings=filter_column_settings,
            include_diagnostics=False,
        )
    )


@app.post("/api/log-diagnostics")
async def log_diagnostics(request: Request):
    form = await request.form()
    raw_case_id_column = form.get("case_id_column")
    raw_activity_column = form.get("activity_column")
    raw_timestamp_column = form.get("timestamp_column")
    filter_column_settings = get_form_filter_column_settings(form)
    file_source, source_file_name = resolve_profile_file_source(form)

    try:
        raw_df = read_raw_log_dataframe(file_source)
    except ValueError as exc:
        return JSONResponse(status_code=400, content={"error": str(exc)})
    except Exception as exc:
        return JSONResponse(
            status_code=400,
            content={
                "error": "Log diagnostics could not be read. Please check the file encoding and header row.",
                "detail": str(exc),
            },
        )

    return JSONResponse(
        content=build_log_profile_payload(
            raw_df=raw_df,
            source_file_name=source_file_name,
            case_id_column=str(raw_case_id_column or "").strip(),
            activity_column=str(raw_activity_column or "").strip(),
            timestamp_column=str(raw_timestamp_column or "").strip(),
            filter_column_settings=filter_column_settings,
            include_diagnostics=True,
        )
    )


@app.post("/api/analyze")
async def analyze(request: Request):
    form = await request.form()
    uploaded_file = form.get("csv_file")
    raw_case_id_column = form.get("case_id_column")
    raw_activity_column = form.get("activity_column")
    raw_timestamp_column = form.get("timestamp_column")
    case_id_column = (
        DEFAULT_HEADERS["case_id_column"]
        if raw_case_id_column is None
        else str(raw_case_id_column).strip()
    )
    activity_column = (
        DEFAULT_HEADERS["activity_column"]
        if raw_activity_column is None
        else str(raw_activity_column).strip()
    )
    timestamp_column = (
        DEFAULT_HEADERS["timestamp_column"]
        if raw_timestamp_column is None
        else str(raw_timestamp_column).strip()
    )
    selected_analysis_keys = form.getlist("analysis_keys")
    filter_column_settings = get_form_filter_column_settings(form)
    base_filter_params = get_form_filter_params(form)

    if uploaded_file and uploaded_file.filename:
        uploaded_file.file.seek(0)
        file_source = uploaded_file.file
        source_file_name = uploaded_file.filename
    else:
        file_source = SAMPLE_FILE
        source_file_name = SAMPLE_FILE.name

    try:
        validate_selected_columns(
            case_id_column=case_id_column,
            activity_column=activity_column,
            timestamp_column=timestamp_column,
        )
        validate_filter_column_settings(filter_column_settings)
        prepared_df = load_prepared_event_log(
            file_source=file_source,
            case_id_column=case_id_column,
            activity_column=activity_column,
            timestamp_column=timestamp_column,
        )
        filtered_prepared_df = filter_prepared_df(
            prepared_df,
            base_filter_params,
            filter_column_settings=filter_column_settings,
        )
        result = analyze_prepared_event_log(
            prepared_df=filtered_prepared_df,
            selected_analysis_keys=selected_analysis_keys,
            output_root_dir=None,
            export_excel=False,
        )
        run_id = save_run_data(
            source_file_name=source_file_name,
            selected_analysis_keys=selected_analysis_keys,
            prepared_df=prepared_df,
            result=result,
            column_settings={
                "case_id_column": case_id_column,
                "activity_column": activity_column,
                "timestamp_column": timestamp_column,
                **filter_column_settings,
            },
            base_filter_params=base_filter_params,
        )
    except ValueError as exc:
        return JSONResponse(status_code=400, content={"error": str(exc)})
    except Exception as exc:
        return JSONResponse(
            status_code=500,
            content={
                "error": "Analysis failed unexpectedly.",
                "detail": str(exc),
            },
        )

    return JSONResponse(
        content=build_preview_response(
            run_id=run_id,
            source_file_name=source_file_name,
            selected_analysis_keys=selected_analysis_keys,
            result=result,
            run_data=get_run_data(run_id),
        )
    )


if __name__ == "__main__":
    uvicorn.run("web_app:app", host="127.0.0.1", port=5000, reload=True)